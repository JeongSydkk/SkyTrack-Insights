# analytics.py
import os
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine, text
import plotly.express as px
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import load_workbook

# =========================
# 1) Подключение к БД
# =========================
# >>> ЗАМЕНИ ПАРОЛЬ <<<
DB_URL = "postgresql+psycopg2://postgres:NewStrongPass123@localhost:5432/otp_analysis"
engine = create_engine(DB_URL, future=True)

# Создадим папки
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

# Немного единого стиля для графиков
plt.rcParams.update({
    "figure.figsize": (10, 6),
    "axes.grid": True,
    "font.size": 11
})

def q(sql: str, params=None) -> pd.DataFrame:
    """Утилита: выполнить SQL и вернуть DataFrame."""
    return pd.read_sql(sql, engine, params=params)

def save_and_report(fig_path: str, df: pd.DataFrame, title: str, note: str):
    """Общий отчёт в консоль для графика."""
    print(f"[OK] Saved chart: {fig_path}")
    print(f"     Rows: {len(df)}")
    print(f"     Title: {title}")
    print(f"     Shows: {note}\n")

# =========================
# 2) Шесть графиков (каждый — с >=2 JOIN)
# =========================

def chart_pie_airline_share():
    """
    Pie: доля выполненных рейсов по авиалиниям (сумма sectors_flown).
    SQL: facts_otp JOIN airlines JOIN calendar_months
    """
    sql = """
    SELECT a.airline_name,
           SUM(f.sectors_flown) AS flights
    FROM facts_otp f
    JOIN airlines a ON f.airline_id = a.airline_id
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY a.airline_name
    HAVING SUM(f.sectors_flown) > 0
    ORDER BY flights DESC;
    """
    df = q(sql)
    # Фильтруем "All Airlines" если есть
    df = df[df["airline_name"].str.lower() != "all airlines"]
    plt.figure()
    plt.title("Share of flights by airline")
    plt.pie(df["flights"], labels=df["airline_name"], autopct="%1.1f%%")
    path = "charts/pie_airline_share.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    save_and_report(path, df, "Share of flights by airline", "Distribution of flown sectors across airlines")

def chart_bar_cancellation_rate_by_airline():
    """
    Bar: процент отмен по авиалиниям (cancellations / scheduled).
    SQL: facts_otp JOIN airlines JOIN calendar_months
    """
    sql = """
    SELECT a.airline_name,
           SUM(f.cancellations) AS cancels,
           SUM(f.sectors_scheduled) AS scheduled
    FROM facts_otp f
    JOIN airlines a ON f.airline_id = a.airline_id
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY a.airline_name
    HAVING SUM(f.sectors_scheduled) > 0;
    """
    df = q(sql)
    df["cancel_rate_pct"] = (df["cancels"] / df["scheduled"] * 100).round(2)
    df = df.sort_values("cancel_rate_pct", ascending=False)
    plt.figure()
    plt.title("Cancellation rate by airline (%)")
    plt.bar(df["airline_name"], df["cancel_rate_pct"])
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("% cancellations")
    path = "charts/bar_cancellation_rate_by_airline.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    save_and_report(path, df, "Cancellation rate by airline", "Cancellations vs scheduled flights (percentage)")

def chart_barh_top10_routes_by_delay_pct():
    """
    Horizontal bar: ТОП-10 маршрутов по доле задержек.
    SQL: facts_otp JOIN routes JOIN ports (2 раза) JOIN calendar_months
    """
    sql = """
    SELECT r.route_id,
           p1.port_name AS origin,
           p2.port_name AS destination,
           SUM(f.departures_delayed + f.arrivals_delayed) AS delayed,
           SUM(f.sectors_flown) AS flown
    FROM facts_otp f
    JOIN routes r ON f.route_id = r.route_id
    JOIN ports p1 ON r.origin_port_id = p1.port_id
    JOIN ports p2 ON r.dest_port_id = p2.port_id
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY r.route_id, origin, destination
    HAVING SUM(f.sectors_flown) > 0;
    """
    df = q(sql)
    df["delay_pct"] = (df["delayed"] / df["flown"] * 100).round(2)
    top10 = df.sort_values("delay_pct", ascending=False).head(10).iloc[::-1]  # перевернем для barh
    plt.figure()
    plt.title("Top-10 routes by delay percentage")
    plt.barh(top10["origin"] + " → " + top10["destination"], top10["delay_pct"])
    plt.xlabel("% delayed")
    path = "charts/barh_top10_routes_by_delay_pct.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    save_and_report(path, top10, "Top-10 routes by delay %", "Routes with the highest delay share")

def chart_line_monthly_ontime_rate_total():
    """
    Line: помесячный тренд punctuality (вылеты, % on-time).
    SQL: facts_otp JOIN calendar_months
    """
    sql = """
    SELECT c.year, c.month_num, c.month_label,
           SUM(f.departures_on_time) AS ontime,
           SUM(f.sectors_flown) AS flown
    FROM facts_otp f
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY c.year, c.month_num, c.month_label
    ORDER BY c.year, c.month_num;
    """
    df = q(sql)
    df = df[df["flown"] > 0]
    df["pct_ontime"] = (df["ontime"] / df["flown"] * 100).round(2)
    df["date"] = pd.to_datetime(dict(year=df.year, month=df.month_num, day=1))
    plt.figure()
    plt.title("Monthly on-time departure rate (%)")
    plt.plot(df["date"], df["pct_ontime"], marker="o")
    plt.ylabel("% on-time")
    plt.xlabel("Month")
    path = "charts/line_monthly_ontime_rate_total.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    save_and_report(path, df, "Monthly on-time rate", "Trend of on-time departures across months")

def chart_hist_delay_rate_distribution():
    """
    Histogram: распределение помесячных delay-rate по маршрутам.
    SQL: facts_otp JOIN routes JOIN ports(2) JOIN calendar_months
    """
    sql = """
    SELECT r.route_id,
           p1.port_name AS origin,
           p2.port_name AS destination,
           c.year, c.month_num,
           SUM(f.departures_delayed + f.arrivals_delayed) AS delayed,
           SUM(f.sectors_flown) AS flown
    FROM facts_otp f
    JOIN routes r ON f.route_id = r.route_id
    JOIN ports p1 ON r.origin_port_id = p1.port_id
    JOIN ports p2 ON r.dest_port_id = p2.port_id
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY r.route_id, origin, destination, c.year, c.month_num;
    """
    df = q(sql)
    df = df[df["flown"] > 0]
    df["delay_rate"] = (df["delayed"] / df["flown"] * 100)
    plt.figure()
    plt.title("Distribution of route-month delay rates (%)")
    df["delay_rate"].plot.hist(bins=30)
    plt.xlabel("% delayed")
    path = "charts/hist_delay_rate_distribution.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    save_and_report(path, df, "Delay rate distribution", "Histogram of monthly delay % across routes")

def chart_scatter_flights_vs_delay_by_airline_month():
    """
    Scatter: связь объёма рейсов и доли задержек по авиалиниям (месячные точки).
    SQL: facts_otp JOIN airlines JOIN calendar_months
    """
    sql = """
    SELECT a.airline_name,
           c.year, c.month_num,
           SUM(f.sectors_flown) AS flown,
           SUM(f.departures_delayed + f.arrivals_delayed) AS delayed
    FROM facts_otp f
    JOIN airlines a ON f.airline_id = a.airline_id
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY a.airline_name, c.year, c.month_num;
    """
    df = q(sql)
    df = df[df["flown"] > 0]
    df["delay_pct"] = (df["delayed"] / df["flown"] * 100)
    plt.figure()
    plt.title("Flights volume vs delay % by airline (monthly points)")
    for name, g in df.groupby("airline_name"):
        plt.scatter(g["flown"], g["delay_pct"], label=name, alpha=0.6, s=25)
    plt.xlabel("Sectors flown (per month)")
    plt.ylabel("% delayed")
    plt.legend(loc="best", fontsize=8, ncol=2)
    path = "charts/scatter_flights_vs_delay_by_airline_month.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    save_and_report(path, df, "Flights vs delay %", "Scatter of monthly flights vs delay % by airline")

# =========================
# 3) Plotly: интерактив с тайм-слайдером
# =========================
def plotly_time_slider():
    """
    Интерактив: объём рейсов по авиалиниям с покадровой анимацией по месяцам.
    """
    sql = """
    SELECT a.airline_name,
           MAKE_DATE(c.year, c.month_num, 1) AS month_date,
           SUM(f.sectors_flown) AS flown
    FROM facts_otp f
    JOIN airlines a ON f.airline_id = a.airline_id
    JOIN calendar_months c ON f.cal_id = c.cal_id
    GROUP BY a.airline_name, month_date
    ORDER BY month_date;
    """
    df = q(sql)
    df["month_date"] = pd.to_datetime(df["month_date"])
    df["month_str"] = df["month_date"].dt.strftime("%Y-%m")
    fig = px.bar(
        df, x="airline_name", y="flown", color="airline_name",
        animation_frame="month_str", title="Sectors flown by airline (time slider)"
    )
    # Сохраняем график и открываем в браузере
    fig.write_html("charts/plotly_slider.html", auto_open=True)
  


def export_to_excel(dataframes: dict, filename: str):
    path = os.path.join("exports", filename)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        total_rows = 0
        for sheet, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            total_rows += len(df)

    # Открываем для форматирования
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Freeze header
        ws.freeze_panes = "B2"
        # Автофильтры
        ws.auto_filter.ref = ws.dimensions
        # Ширина колонок
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            header = col[0].value
            width = max(10, len(str(header)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Градиент для числовых колонок
        if ws.max_column >= 2 and ws.max_row >= 2:
            # Найдём числовые столбцы по первой «данной» строке
            numeric_cols = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=2, column=col_idx).value
                if isinstance(cell, (int, float)):
                    numeric_cols.append(col_idx)
            for col_idx in numeric_cols:
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(rng, rule)
    wb.save(path)

    sheets_count = len(dataframes)
    rows_count = sum(len(df) for df in dataframes.values())
    print(f'[OK] Created file {filename}, {sheets_count} sheets, {rows_count} rows -> exports/{filename}')

# =========================
# 5) Демо-вставка строки и регенерация графика
# =========================
def insert_demo_row_and_refresh():
    """
    Для защиты: вставим 1 строку в facts_otp (маленькие числа),
    затем перегенерируем линейный график, чтобы он изменился.
    """
    # Берём любую существующую комбинацию airline/route/period
    pick_sql = """
    SELECT f.route_id, f.airline_id, f.cal_id
    FROM facts_otp f
    LIMIT 1;
    """
    pick = q(pick_sql)
    if pick.empty:
        print("[WARN] No data in facts_otp to demo-insert.")
        return
    route_id = int(pick.loc[0, "route_id"])
    airline_id = int(pick.loc[0, "airline_id"])
    cal_id = int(pick.loc[0, "cal_id"])

    # Вставляем крошечную запись (чтобы заметно сдвинуть метрику)
    insert_sql = text("""
        INSERT INTO facts_otp (
            route_id, airline_id, cal_id,
            sectors_scheduled, sectors_flown, cancellations,
            departures_on_time, arrivals_on_time,
            departures_delayed, arrivals_delayed
        )
        VALUES (:route_id, :airline_id, :cal_id,
                2, 2, 0,
                1, 1,
                1, 1);
    """)
    with engine.begin() as conn:
        conn.execute(insert_sql, {
            "route_id": route_id,
            "airline_id": airline_id,
            "cal_id": cal_id
        })
    print("[OK] Inserted 1 demo row into facts_otp.")
    # Регенерируем один график
    chart_line_monthly_ontime_rate_total()
    print("[OK] Regenerated line chart after demo insert.\n")

# =========================
# 6) Сборка и запуск
# =========================
def main():
    print("=== Building 6 charts ===")
    chart_pie_airline_share()
    chart_bar_cancellation_rate_by_airline()
    chart_barh_top10_routes_by_delay_pct()
    chart_line_monthly_ontime_rate_total()
    chart_hist_delay_rate_distribution()
    chart_scatter_flights_vs_delay_by_airline_month()

    print("=== Plotly (time slider) ===")
    plotly_time_slider()

    print("=== Export to Excel ===")
    # Подготовим несколько таблиц для экспорта
    df_airline_share = q("""
        SELECT a.airline_name, SUM(f.sectors_flown) AS flights
        FROM facts_otp f
        JOIN airlines a ON f.airline_id = a.airline_id
        JOIN calendar_months c ON f.cal_id = c.cal_id
        GROUP BY a.airline_name
        ORDER BY flights DESC;""")

    df_top_routes = q("""
        SELECT p1.port_name AS origin, p2.port_name AS destination,
               SUM(f.departures_delayed + f.arrivals_delayed) AS delayed,
               SUM(f.sectors_flown) AS flown
        FROM facts_otp f
        JOIN routes r ON f.route_id = r.route_id
        JOIN ports p1 ON r.origin_port_id = p1.port_id
        JOIN ports p2 ON r.dest_port_id = p2.port_id
        JOIN calendar_months c ON f.cal_id = c.cal_id
        GROUP BY origin, destination
        HAVING SUM(f.sectors_flown) > 0
        ORDER BY delayed DESC
        LIMIT 25;""")

    df_monthly = q("""
        SELECT c.year, c.month_num,
               SUM(f.departures_on_time) AS ontime,
               SUM(f.sectors_flown) AS flown
        FROM facts_otp f
        JOIN calendar_months c ON f.cal_id = c.cal_id
        GROUP BY c.year, c.month_num
        ORDER BY c.year, c.month_num;""")

    export_to_excel({
        "AirlineShare": df_airline_share,
        "TopRoutesDelays": df_top_routes,
        "MonthlyOnTime": df_monthly
    }, filename=f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

    print("=== Demo insert & refresh ===")
    insert_demo_row_and_refresh()

    print("\nAll done.")

if __name__ == "__main__":
    main()
